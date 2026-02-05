from openpyxl import load_workbook
import csv

wb = load_workbook(r'C:\Users\mamingli\Desktop\项目资料\大贺运输新系统\別紙_画面一覧_v151.xlsx')
ws = wb.worksheets[25]

with open(r'C:\Users\mamingli\Desktop\项目资料\大贺运输新系统\別紙_画面一覧_v151\p1-スタート\data_raw.csv', 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(row)

print("Data extracted to data_raw.csv")
